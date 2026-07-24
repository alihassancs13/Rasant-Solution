"""Admin worklog analytics — local DB (Jira-synced + manual), 8h = 1 day."""
from __future__ import annotations

from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone
from io import BytesIO
import calendar
import re

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.hyperlink import Hyperlink

from django.contrib.auth import get_user_model
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import JiraCredential, Worklog, WorklogExportSettings
from .worklog_storage import (
    seconds_to_day_units,
    seconds_to_hours,
    serialize_worklog_row,
    sync_user_worklogs_from_jira,
    worklogs_in_range,
)


HOURS_PER_DAY = Worklog.HOURS_PER_DAY
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F2"),
    right=Side(style="thin", color="D9E2F2"),
    top=Side(style="thin", color="D9E2F2"),
    bottom=Side(style="thin", color="D9E2F2"),
)


def _is_admin_user(user):
    role_name = (user.role.name if getattr(user, "role", None) else "") or ""
    return role_name.lower() in ("admin", "administrator") or user.is_superuser or user.is_staff


def _parse_analytics_range(request):
    today = datetime.now(timezone.utc).date()
    default_from = today.replace(day=1)
    date_from_raw = request.query_params.get("from") or request.query_params.get("date_from")
    date_to_raw = request.query_params.get("to") or request.query_params.get("date_to")
    try:
        date_from = datetime.strptime(date_from_raw, "%Y-%m-%d").date() if date_from_raw else default_from
    except ValueError:
        date_from = default_from
    try:
        date_to = datetime.strptime(date_to_raw, "%Y-%m-%d").date() if date_to_raw else today
    except ValueError:
        date_to = today
    if date_to < date_from:
        date_from, date_to = date_to, date_from
    return date_from, date_to


def _month_bounds_for_export(request):
    today = datetime.now(timezone.utc).date()
    month_raw = request.query_params.get("month")
    ref_raw = request.query_params.get("from") or request.query_params.get("date_from")
    target = today
    if month_raw:
        try:
            target = datetime.strptime(month_raw, "%Y-%m").date()
        except ValueError:
            target = today
    elif ref_raw:
        try:
            target = datetime.strptime(ref_raw, "%Y-%m-%d").date()
        except ValueError:
            target = today
    first = target.replace(day=1)
    last = target.replace(day=calendar.monthrange(target.year, target.month)[1])
    return first, last


def _owner_from_user(user, employee=None, cred=None):
    name = None
    email = (cred.email if cred else None) or (user.email if user else "")
    if employee:
        name = employee.name or name
        email = employee.email or email
    if user and not name:
        name = f"{user.first_name} {user.last_name}".strip() or user.username
    return {
        "user_id": user.id if user else None,
        "employee_id": employee.id if employee else None,
        "name": name or email or "Unknown",
        "email": email or "",
        "department": employee.department if employee else "",
        "designation": employee.designation if employee else "",
        "is_active": employee.is_active if employee else (user.is_active if user else True),
    }


def _sync_all_jira_for_range(date_from, date_to):
    """Best-effort sync of every connected Jira account into the local table."""
    creds = list(
        JiraCredential.objects.select_related("auth_user_id")
        .exclude(domain__isnull=True)
        .exclude(domain="")
        .exclude(api_token__isnull=True)
        .exclude(api_token="")
    )
    errors = []

    def _one(cred):
        from django.db import close_old_connections
        close_old_connections()
        user = cred.auth_user_id
        if not user:
            return None
        count, err = sync_user_worklogs_from_jira(user, date_from, date_to)
        return {"user_id": user.id, "synced": count, "error": err}

    max_workers = min(8, max(len(creds), 1))
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = [pool.submit(_one, c) for c in creds]
        for fut in as_completed(futures):
            try:
                result = fut.result()
            except Exception as exc:
                errors.append({"error": str(exc)})
                continue
            if result and result.get("error"):
                errors.append(result)
    return errors


def _week_bucket(day: date) -> str:
    if day.day <= 7:
        return "W1"
    if day.day <= 14:
        return "W2"
    if day.day <= 21:
        return "W3"
    if day.day <= 28:
        return "W4"
    return "W5"


def _hours(value) -> float:
    return round((int(value or 0) / 3600), 3)


def _clip(text, limit=60):
    value = " ".join(str(text or "").split())
    if len(value) <= limit:
        return value
    return value[: max(limit - 1, 0)].rstrip() + "…"


def _autosize(ws, widths: dict[str, int]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_cell(cell, *, fill=None, bold=False, color="1F2937", align="left", size=10):
    cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _fetch_jira_display_name(cred: JiraCredential) -> str:
    """Best-effort fetch of Jira displayName; persists when found."""
    if not cred or not cred.domain or not cred.email or not cred.api_token:
        return (cred.display_name if cred else "") or ""
    if cred.display_name:
        return cred.display_name
    try:
        import requests
        from .services import get_jira_headers

        domain, headers = get_jira_headers(cred.email, cred.api_token, cred.domain)
        r = requests.get(f"{domain}/rest/api/3/myself", headers=headers, timeout=8)
        if r.status_code == 200:
            name = (r.json().get("displayName") or "").strip()
            if name:
                JiraCredential.objects.filter(pk=cred.pk).update(display_name=name)
                cred.display_name = name
                return name
    except Exception:
        pass
    return ""


def _resolve_export_names(rows):
    """
    Map user_id -> Jira display name (preferred) for Responsible / Employee columns.
    Falls back to employee name / system name / email.
    """
    from employeeDashboard.models import Employee

    user_ids = {r.user_id for r in rows if r.user_id}
    employees = {
        e.user_id: e
        for e in Employee.objects.filter(user_id__in=user_ids)
    }
    creds = {
        c.auth_user_id_id: c
        for c in JiraCredential.objects.filter(auth_user_id_id__in=user_ids)
    }

    # Refresh missing Jira display names in parallel
    missing = [c for c in creds.values() if not (c.display_name or "").strip()]
    if missing:
        with ThreadPoolExecutor(max_workers=min(8, max(len(missing), 1))) as pool:
            list(pool.map(_fetch_jira_display_name, missing))

    names = {}
    for uid in user_ids:
        cred = creds.get(uid)
        emp = employees.get(uid)
        jira_name = (cred.display_name if cred else "") or ""
        emp_name = (emp.name if emp else "") or ""
        # Prefer Jira username/display name for the sheet
        names[uid] = jira_name or emp_name or (cred.email if cred else "") or "Unknown"
    return names


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet titles: max 31 chars, no \\ / ? * [ ]."""
    cleaned = re.sub(r'[\\/*?:\[\]]', " ", str(name or "Employee")).strip() or "Employee"
    cleaned = " ".join(cleaned.split())
    base = cleaned[:28]
    title = base
    n = 2
    while title in used:
        suffix = f" ({n})"
        title = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(title)
    return title


def _quote_sheet(title: str) -> str:
    return "'" + str(title).replace("'", "''") + "'"


def _configure_week_hours_axes(chart):
    """Proper X (Week) / Y (Hours 0–40 step 5) axis labels and ticks."""
    chart.x_axis.title = "Week"
    chart.y_axis.title = "Hours"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.minorTickMark = "none"
    chart.y_axis.minorTickMark = "none"
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 40
    chart.y_axis.majorUnit = 5
    chart.y_axis.minorUnit = 5
    chart.y_axis.numFmt = "0"
    chart.y_axis.majorGridlines = ChartLines()
    # Leave space so axis titles render clearly
    chart.width = 15
    chart.height = 10


def _add_weekly_bar_chart(
    ws,
    *,
    title,
    anchor,
    color="4A90E2",
    # Vertical source (employee sheets): Week | Hours columns
    data_col=None,
    cat_col=None,
    header_row=None,
    last_data_row=None,
    # Horizontal source (Summary): W1–W4 header row + values row
    first_col=None,
    last_col=None,
    cats_row=None,
    values_row=None,
):
    """
    Column chart: X = W1..W4, Y = hours 0–40 (step 5).
    Hover in Excel shows category (week) + value (hours).
    Data must be in visible cells — Excel often skips charts bound to hidden columns.
    """
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.legend = None
    _configure_week_hours_axes(chart)

    if cats_row is not None and values_row is not None and first_col is not None and last_col is not None:
        # One series across columns (Summary project W1–W4)
        cats_ref = Reference(ws, min_col=first_col, max_col=last_col, min_row=cats_row, max_row=cats_row)
        data_ref = Reference(ws, min_col=first_col, max_col=last_col, min_row=values_row, max_row=values_row)
        chart.add_data(data_ref, from_rows=True, titles_from_data=False)
        chart.set_categories(cats_ref)
    else:
        # Vertical Week / Hours table (include header so series is named "Hours")
        data_ref = Reference(ws, min_col=data_col, min_row=header_row, max_row=last_data_row)
        cats_ref = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

    try:
        if chart.series:
            series = chart.series[0]
            series.graphicalProperties.solidFill = color
            # Explicit series name → Excel hover: Week + Hours
            series.tx = SeriesLabel(v="Hours")
            # Force string category refs so W1/W2/... show on the X axis
            if series.cat is not None:
                formula = None
                if series.cat.numRef is not None:
                    formula = series.cat.numRef.f
                elif series.cat.strRef is not None:
                    formula = series.cat.strRef.f
                if formula:
                    series.cat.numRef = None
                    series.cat.strRef = StrRef(f=formula)
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.showCatName = False
            series.dLbls.showSerName = False
    except Exception:
        pass

    ws.add_chart(chart, anchor)
    return chart


def _build_worklog_export_workbook(*, settings_obj, rows, date_from, date_to):
    """
    Dynamic workbook:
      Summary — editable project name/no; employee W1–W5 via formulas to each sheet
      Employee sheets — editable hours; project fields + weekly stats + chart via formulas
    """
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    month_label = date_from.strftime("%B %Y")
    project_name = settings_obj.project_name or "Project"
    project_number = settings_obj.project_number or "—"
    # Chart / analytics weeks (W1–W4 primary axis labels; W5 kept in tables)
    week_keys = ["W1", "W2", "W3", "W4", "W5"]
    chart_weeks = ["W1", "W2", "W3", "W4"]  # X-axis as requested
    name_by_user = _resolve_export_names(rows)

    employee_week = defaultdict(lambda: {k: 0.0 for k in week_keys + ["MD"]})
    employee_day = defaultdict(lambda: defaultdict(float))
    employee_details = defaultdict(list)
    employee_tickets = defaultdict(set)

    for row in rows:
        user = getattr(row, "user", None)
        started = getattr(row, "started", None)
        day = started.date() if started else None
        if not user or not day:
            continue

        employee_name = name_by_user.get(user.id) or (
            f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip()
            or getattr(user, "username", "")
            or getattr(user, "email", "")
            or "Unknown"
        )
        hours = _hours(getattr(row, "time_spent_seconds", 0))
        week_key = _week_bucket(day)
        issue_key = (getattr(row, "issue_key", None) or "").strip()
        summary = getattr(row, "summary", None) or ""
        comment = getattr(row, "comment", None) or ""

        employee_week[employee_name][week_key] += hours
        employee_week[employee_name]["MD"] += hours
        employee_day[employee_name][day.day] += hours
        if issue_key:
            employee_tickets[employee_name].add(issue_key)

        employee_details[employee_name].append(
            {
                "date": day,
                "date_label": day.strftime("%d-%b-%Y"),
                "hours": hours,
                "ticket_no": issue_key or "—",
                "ticket_description": _clip(summary, 100) or "—",
                "details": _clip(comment, 100) or "—",
                "responsible": employee_name,
                "week": week_key,
            }
        )

    employee_names = sorted(employee_week.keys())
    used_titles: set[str] = {"Summary"}
    sheet_title_by_name = {name: _safe_sheet_title(name, used_titles) for name in employee_names}

    # Fixed cells on each employee sheet for Summary formulas:
    # K5..K9 = W1..W5 hours, K10 = MD  (analytics table)
    EMP_WEEK_ROW = {wk: 5 + i for i, wk in enumerate(week_keys)}  # W1->5 ... W5->9
    EMP_MD_ROW = 10
    EMP_HOURS_COL = 11  # K

    # ─── Summary ─────────────────────────────────────────────────────
    summary_ws.merge_cells("A1:H1")
    summary_ws["A1"] = f"Timesheet Summary — {month_label}"
    summary_ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    summary_ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    summary_ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    summary_ws.row_dimensions[1].height = 24

    summary_ws.merge_cells("A2:H2")
    summary_ws["A2"] = (
        "Edit Project / Project No. below — employee sheets update automatically. "
        "Click an employee name to open their timesheet."
    )
    summary_ws["A2"].font = Font(bold=True, size=10, color="1E3A5F", name="Calibri")
    summary_ws["A2"].fill = PatternFill("solid", fgColor="EAF3FF")
    summary_ws.row_dimensions[2].height = 18

    for idx, label in enumerate(["Project", "Project No."] + week_keys + ["MD"], start=1):
        _style_cell(summary_ws.cell(4, idx, label), fill="DCEBFF", bold=True, color="1E3A5F", align="center")

    # Editable project identity
    _style_cell(summary_ws.cell(5, 1, project_name), fill="FFFDE7", bold=True, align="left")
    _style_cell(summary_ws.cell(5, 2, project_number), fill="FFFDE7", bold=True, align="left")

    first_emp_row = 9
    last_emp_row = first_emp_row + max(len(employee_names) - 1, 0) if employee_names else first_emp_row

    # Project week totals = sum of employee formula columns (dynamic)
    for col_idx, _wk in enumerate(week_keys, start=3):
        col_letter = chr(ord("A") + col_idx - 1)
        if employee_names:
            formula = f"=SUM({col_letter}{first_emp_row}:{col_letter}{last_emp_row})"
        else:
            formula = 0
        _style_cell(summary_ws.cell(5, col_idx, formula), fill="F8FBFF", align="center")
    if employee_names:
        md_formula = f"=SUM(H{first_emp_row}:H{last_emp_row})"
    else:
        md_formula = 0
    _style_cell(summary_ws.cell(5, 8, md_formula), fill="F8FBFF", bold=True, align="center")

    for idx in range(1, 9):
        cell = summary_ws.cell(6, idx)
        if idx == 1:
            cell.value = "Total"
        elif idx >= 3:
            # Mirror project row so Total stays in sync
            col_letter = chr(ord("A") + idx - 1)
            cell.value = f"={col_letter}5"
        _style_cell(cell, fill="E8F3E8", bold=True, color="14532D", align="center" if idx >= 3 else "left")

    for idx, label in enumerate(["Employee"] + week_keys + ["MD"], start=1):
        _style_cell(summary_ws.cell(8, idx, label), fill="DCEBFF", bold=True, color="1E3A5F", align="center")

    for i, name in enumerate(employee_names):
        r = first_emp_row + i
        title = sheet_title_by_name[name]
        q = _quote_sheet(title)
        fill = "FFFFFF" if r % 2 else "F8FBFF"

        name_cell = summary_ws.cell(r, 1, name)
        _style_cell(name_cell, fill=fill, align="left")
        name_cell.hyperlink = Hyperlink(ref=name_cell.coordinate, location=f"{q}!A1")
        name_cell.font = Font(bold=True, color="0563C1", underline="single", size=10, name="Calibri")

        for col_idx, wk in enumerate(week_keys, start=3):
            week_row = EMP_WEEK_ROW[wk]
            formula = f"={q}!{chr(64 + EMP_HOURS_COL)}{week_row}"
            _style_cell(summary_ws.cell(r, col_idx, formula), fill=fill, align="center")
        md_f = f"={q}!{chr(64 + EMP_HOURS_COL)}{EMP_MD_ROW}"
        _style_cell(summary_ws.cell(r, 8, md_f), fill=fill, bold=True, align="center")
        summary_ws.row_dimensions[r].height = 17

    total_row = last_emp_row + 1 if employee_names else first_emp_row
    if employee_names:
        for idx in range(1, 9):
            cell = summary_ws.cell(total_row, idx)
            if idx == 1:
                cell.value = "Total"
            elif idx >= 3:
                col_letter = chr(ord("A") + idx - 1)
                cell.value = f"=SUM({col_letter}{first_emp_row}:{col_letter}{last_emp_row})"
            _style_cell(cell, fill="E8F3E8", bold=True, color="14532D", align="center" if idx >= 2 else "left")

    # Chart uses visible Summary W1–W4 headers (C4:F4) + project totals (C5:F5).
    # Do not bind charts to hidden columns — Excel often fails to render them.
    _add_weekly_bar_chart(
        summary_ws,
        title=f"Project Weekly Hours — {month_label}",
        first_col=3,
        last_col=6,
        cats_row=4,
        values_row=5,
        anchor="J4",
        color="4A90E2",
    )

    _autosize(summary_ws, {"A": 28, "B": 18, "C": 9, "D": 9, "E": 9, "F": 9, "G": 9, "H": 10})
    summary_ws.freeze_panes = "A4"
    summary_ws.page_setup.orientation = "landscape"
    summary_ws.page_setup.fitToPage = True
    summary_ws.page_setup.fitToWidth = 1
    summary_ws.page_setup.fitToHeight = 1

    detail_headers = [
        "Date",
        "Project",
        "Project No.",
        "No. of hours",
        "Ticket no.",
        "Ticket Description",
        "Details",
        "Responsible",
    ]

    # ─── Per-employee sheets ─────────────────────────────────────────
    for name in employee_names:
        title = sheet_title_by_name[name]
        ws = wb.create_sheet(title)
        details = sorted(
            employee_details.get(name, []),
            key=lambda d: (d["date"], d.get("ticket_no") or ""),
        )

        # Header with live project refs
        ws.merge_cells("A1:H1")
        ws["A1"] = f'="Timesheet Detail — {month_label}  ·  "&Summary!$A$5&" ("&Summary!$B$5&")"'
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:H2")
        back = ws["A2"]
        back.value = "Back to Summary"
        back.hyperlink = Hyperlink(ref="A2", location="'Summary'!A1")
        back.font = Font(bold=True, color="0563C1", underline="single", size=10, name="Calibri")
        back.fill = PatternFill("solid", fgColor="EAF3FF")

        ws.merge_cells("J2:M2")
        ws["J2"] = f"Employee: {name}"
        ws["J2"].font = Font(bold=True, size=10, color="1E3A5F", name="Calibri")
        ws["J2"].fill = PatternFill("solid", fgColor="EAF3FF")

        ws.merge_cells("A3:H3")
        ws["A3"] = "Timesheet Entries  ·  Edit hours in column D — totals & chart update automatically"
        ws["A3"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        ws["A3"].fill = PatternFill("solid", fgColor="1E3A5F")
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

        header_row = 4
        for idx, label in enumerate(detail_headers, start=1):
            _style_cell(
                ws.cell(header_row, idx, label),
                fill="DCEBFF",
                bold=True,
                color="1E3A5F",
                align="center",
                size=9,
            )

        detail_start = header_row + 1
        if not details:
            _style_cell(
                ws.cell(detail_start, 1, "No worklog entries for this employee in the selected month."),
                fill="FFF7ED",
                color="9A3412",
                align="left",
                size=9,
            )
            ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=8)
            last_data_row = detail_start
            total_row = detail_start + 1
        else:
            for row_idx, payload in enumerate(details, start=detail_start):
                fill = "FFFFFF" if row_idx % 2 else "F8FBFF"
                _style_cell(ws.cell(row_idx, 1, payload["date_label"]), fill=fill, align="center", size=9)
                # Live project identity from Summary
                _style_cell(ws.cell(row_idx, 2, "=Summary!$A$5"), fill=fill, align="left", size=9)
                _style_cell(ws.cell(row_idx, 3, "=Summary!$B$5"), fill=fill, align="left", size=9)
                # Editable hours (highlighted)
                hours_cell = ws.cell(row_idx, 4, payload["hours"])
                _style_cell(hours_cell, fill="FFFDE7", align="center", size=9)
                _style_cell(ws.cell(row_idx, 5, payload["ticket_no"]), fill=fill, align="left", size=9)
                _style_cell(ws.cell(row_idx, 6, payload["ticket_description"]), fill=fill, align="left", size=9)
                _style_cell(ws.cell(row_idx, 7, payload["details"]), fill=fill, align="left", size=9)
                _style_cell(ws.cell(row_idx, 8, payload["responsible"]), fill=fill, align="left", size=9)
                # Hidden week key for SUMIF/COUNTIF
                ws.cell(row_idx, 9, payload["week"])
                ws.row_dimensions[row_idx].height = 15
            last_data_row = detail_start + len(details) - 1
            total_row = last_data_row + 1

        for idx in range(1, 9):
            cell = ws.cell(total_row, idx)
            if idx == 1:
                cell.value = "Total"
            elif idx == 4:
                if details:
                    cell.value = f"=SUM(D{detail_start}:D{last_data_row})"
                else:
                    cell.value = 0
            _style_cell(cell, fill="E8F3E8", bold=True, color="14532D", align="center" if idx == 4 else "left")

        ws.column_dimensions["I"].hidden = True  # week keys

        # RIGHT — Week / Hours / % / Entries (formulas) + chart  [no meta "Weekly Analytics" block]
        ws.merge_cells("J3:M3")
        ws["J3"] = "Week breakdown"
        ws["J3"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        ws["J3"].fill = PatternFill("solid", fgColor="4A90E2")
        ws["J3"].alignment = Alignment(horizontal="left", vertical="center")

        analytics_header = 4
        for col, label in enumerate(["Week", "Hours", "% of month", "Entries"], start=10):
            _style_cell(
                ws.cell(analytics_header, col, label),
                fill="DCEBFF",
                bold=True,
                color="1E3A5F",
                align="center",
            )

        # Data range for SUMIF — if no rows, use a dummy empty range
        if details:
            week_range = f"$I${detail_start}:$I${last_data_row}"
            hours_range = f"$D${detail_start}:$D${last_data_row}"
        else:
            week_range = "$I$5:$I$5"
            hours_range = "$D$5:$D$5"

        for i, wk in enumerate(week_keys):
            r = analytics_header + 1 + i
            fill = "FFFFFF" if i % 2 == 0 else "F8FBFF"
            _style_cell(ws.cell(r, 10, wk), fill=fill, align="center")
            _style_cell(
                ws.cell(r, 11, f'=SUMIF({week_range},"{wk}",{hours_range})'),
                fill="FFFDE7",
                align="center",
            )
            # % of month vs MD (K10)
            _style_cell(
                ws.cell(r, 12, f"=IF($K$10=0,0,ROUND(K{r}/$K$10*100,1))"),
                fill=fill,
                align="center",
            )
            _style_cell(
                ws.cell(r, 13, f'=COUNTIF({week_range},"{wk}")'),
                fill=fill,
                align="center",
            )

        md_row = analytics_header + 6  # 10
        _style_cell(ws.cell(md_row, 10, "MD"), fill="E8F3E8", bold=True, color="14532D", align="center")
        _style_cell(ws.cell(md_row, 11, "=SUM(K5:K9)"), fill="E8F3E8", bold=True, color="14532D", align="center")
        _style_cell(ws.cell(md_row, 12, 100), fill="E8F3E8", bold=True, color="14532D", align="center")
        _style_cell(ws.cell(md_row, 13, "=SUM(M5:M9)"), fill="E8F3E8", bold=True, color="14532D", align="center")

        # Chart uses W1–W4 only (rows 5–8), Y axis 0–40
        _add_weekly_bar_chart(
            ws,
            title=f"{name} — Weekly Hours",
            data_col=11,
            cat_col=10,
            header_row=analytics_header,
            last_data_row=analytics_header + len(chart_weeks),
            anchor="J12",
            color="1E3A5F",
        )

        _autosize(
            ws,
            {
                "A": 12,
                "B": 14,
                "C": 16,
                "D": 12,
                "E": 12,
                "F": 28,
                "G": 24,
                "H": 18,
                "J": 10,
                "K": 10,
                "L": 12,
                "M": 10,
            },
        )
        ws.freeze_panes = "A5"
        if details:
            ws.auto_filter.ref = f"A{header_row}:H{last_data_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = f"{header_row}:{header_row}"

    return wb


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def worklog_analytics_overview(request):
    if not _is_admin_user(request.user):
        return Response({"success": False, "message": "Admin access required."}, status=403)

    from employeeDashboard.models import Employee

    date_from, date_to = _parse_analytics_range(request)
    fetch_errors = _sync_all_jira_for_range(date_from, date_to)

    User = get_user_model()
    rows = list(worklogs_in_range(date_from, date_to).select_related("user", "jira_credential"))

    user_ids = {r.user_id for r in rows if r.user_id}
    # Also include users who only have credentials (zero after sync still ok to skip)
    employees = {
        e.user_id: e
        for e in Employee.objects.filter(user_id__in=user_ids)
    }
    users = {u.id: u for u in User.objects.filter(id__in=user_ids)}

    by_employee = {}
    daily = defaultdict(lambda: {"seconds": 0, "entries": 0})
    issues = defaultdict(lambda: {"seconds": 0, "entries": 0, "summary": ""})
    total_seconds = 0
    entry_count = 0

    for row in rows:
        user = row.user or users.get(row.user_id)
        if not user:
            continue
        key = user.id
        if key not in by_employee:
            by_employee[key] = {
                **_owner_from_user(user, employees.get(key), row.jira_credential),
                "total_seconds": 0,
                "entries": 0,
                "issues": set(),
                "days": set(),
                "manual_entries": 0,
                "jira_entries": 0,
            }

        bucket = by_employee[key]
        secs = int(row.time_spent_seconds or 0)
        day = row.started.date().isoformat() if row.started else ""
        issue_key = row.issue_key or "—"

        bucket["total_seconds"] += secs
        bucket["entries"] += 1
        bucket["issues"].add(issue_key)
        if day:
            bucket["days"].add(day)
        if row.source == Worklog.SOURCE_MANUAL:
            bucket["manual_entries"] += 1
        else:
            bucket["jira_entries"] += 1

        total_seconds += secs
        entry_count += 1
        if day:
            daily[day]["seconds"] += secs
            daily[day]["entries"] += 1
        issues[issue_key]["seconds"] += secs
        issues[issue_key]["entries"] += 1
        if row.summary:
            issues[issue_key]["summary"] = row.summary

    employees_out = []
    for row in by_employee.values():
        if row["entries"] == 0:
            continue
        hours = seconds_to_hours(row["total_seconds"])
        employees_out.append({
            "user_id": row["user_id"],
            "employee_id": row["employee_id"],
            "name": row["name"],
            "email": row["email"],
            "department": row["department"],
            "designation": row["designation"],
            "is_active": row["is_active"],
            "total_seconds": row["total_seconds"],
            "total_hours": hours,
            "day_units": seconds_to_day_units(row["total_seconds"]),
            "hours_per_day": HOURS_PER_DAY,
            "entries": row["entries"],
            "manual_entries": row["manual_entries"],
            "jira_entries": row["jira_entries"],
            "unique_issues": len(row["issues"]),
            "days_logged": len(row["days"]),
            "days": seconds_to_day_units(row["total_seconds"]),
        })
    employees_out.sort(key=lambda x: x["total_seconds"], reverse=True)

    day_span = max((date_to - date_from).days + 1, 1)
    daily_trend = [
        {
            "date": d,
            "seconds": v["seconds"],
            "hours": seconds_to_hours(v["seconds"]),
            "day_units": seconds_to_day_units(v["seconds"]),
            "entries": v["entries"],
        }
        for d, v in sorted(daily.items())
    ]

    top_issues = sorted(
        [
            {
                "issue_key": k,
                "summary": v["summary"] or "",
                "seconds": v["seconds"],
                "hours": seconds_to_hours(v["seconds"]),
                "entries": v["entries"],
            }
            for k, v in issues.items()
        ],
        key=lambda x: x["seconds"],
        reverse=True,
    )[:10]

    return Response({
        "success": True,
        "data": {
            "range": {"from": date_from.isoformat(), "to": date_to.isoformat()},
            "source": "database",
            "hours_per_day": HOURS_PER_DAY,
            "summary": {
                "total_seconds": total_seconds,
                "total_hours": seconds_to_hours(total_seconds),
                "total_day_units": seconds_to_day_units(total_seconds),
                "total_entries": entry_count,
                "unique_issues": len(issues),
                "active_loggers": len(employees_out),
                "avg_hours_per_day": seconds_to_hours(total_seconds / day_span),
            },
            "by_employee": employees_out,
            "daily_trend": daily_trend,
            "top_issues": top_issues,
            "fetch_errors": fetch_errors[:20],
        },
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def worklog_analytics_employee(request, user_id):
    if not _is_admin_user(request.user):
        return Response({"success": False, "message": "Admin access required."}, status=403)

    from employeeDashboard.models import Employee

    User = get_user_model()
    try:
        target_user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({"success": False, "message": "User not found."}, status=404)

    date_from, date_to = _parse_analytics_range(request)

    # Sync this user's Jira into DB when possible
    sync_error = None
    try:
        _, sync_error = sync_user_worklogs_from_jira(target_user, date_from, date_to)
    except Exception as exc:
        sync_error = str(exc)

    employee = Employee.objects.filter(user=target_user).first()
    cred = JiraCredential.objects.filter(auth_user_id=target_user).first()
    owner = _owner_from_user(target_user, employee, cred)

    rows = list(worklogs_in_range(date_from, date_to, user=target_user))
    total_seconds = sum(int(r.time_spent_seconds or 0) for r in rows)
    days = {r.started.date().isoformat() for r in rows if r.started}
    issues = {r.issue_key for r in rows if r.issue_key}

    entries_out = [serialize_worklog_row(r) for r in rows]

    return Response({
        "success": True,
        "data": {
            "range": {"from": date_from.isoformat(), "to": date_to.isoformat()},
            "source": "database",
            "hours_per_day": HOURS_PER_DAY,
            "employee": owner,
            "summary": {
                "total_seconds": total_seconds,
                "total_hours": seconds_to_hours(total_seconds),
                "day_units": seconds_to_day_units(total_seconds),
                "entries": len(rows),
                "unique_issues": len(issues),
                "days_logged": len(days),
            },
            "entries": entries_out,
            "sync_error": sync_error,
        },
    })


@api_view(["GET", "PUT", "PATCH"])
@permission_classes([IsAuthenticated])
def worklog_export_settings_view(request):
    if not _is_admin_user(request.user):
        return Response({"success": False, "message": "Admin access required."}, status=403)

    settings_obj = WorklogExportSettings.get_solo()
    if request.method == "GET":
        return Response(
            {
                "success": True,
                "data": {
                    "project_name": settings_obj.project_name,
                    "project_number": settings_obj.project_number,
                    "updated_at": settings_obj.updated_at.isoformat() if settings_obj.updated_at else None,
                },
            }
        )

    project_name = (request.data.get("project_name") or "").strip()
    project_number = (request.data.get("project_number") or "").strip()
    if not project_name:
        return Response({"success": False, "message": "Project name is required."}, status=400)
    if not project_number:
        return Response({"success": False, "message": "Project number is required."}, status=400)

    settings_obj.project_name = project_name[:255]
    settings_obj.project_number = project_number[:255]
    settings_obj.save()
    return Response(
        {
            "success": True,
            "message": "Export settings saved.",
            "data": {
                "project_name": settings_obj.project_name,
                "project_number": settings_obj.project_number,
                "updated_at": settings_obj.updated_at.isoformat() if settings_obj.updated_at else None,
            },
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def worklog_analytics_export(request):
    if not _is_admin_user(request.user):
        return Response({"success": False, "message": "Admin access required."}, status=403)

    date_from, date_to = _month_bounds_for_export(request)
    _sync_all_jira_for_range(date_from, date_to)
    rows = list(worklogs_in_range(date_from, date_to).select_related("user", "jira_credential"))
    settings_obj = WorklogExportSettings.get_solo()
    wb = _build_worklog_export_workbook(
        settings_obj=settings_obj,
        rows=rows,
        date_from=date_from,
        date_to=date_to,
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"timesheet_{date_from.strftime('%Y_%m')}.xlsx"
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
